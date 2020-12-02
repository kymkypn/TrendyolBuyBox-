from parsel import Selector
from openpyxl import load_workbook
import re , requests

ac=load_workbook(filename='trendurunler.xlsx')

oku=ac['urunler']

yazAc=load_workbook(filename='trendyolKontrol.xlsx',read_only=False)
yazOku=yazAc['Sayfa1']
yazOku.delete_cols(1)
yazAc.save('trendyolKontrol.xlsx')
yazAc.close()

yazAc=load_workbook(filename='trendyolKontrol.xlsx',read_only=False)
yazOku=yazAc['Sayfa1']
altSira=0

hucreNo=1


while True:

    hucreKontrol = oku['A{}'.format(hucreNo)].value

    if hucreKontrol == None:

        print("Kontrol Bitti")
        hucreurunsay=hucreNo-1
        altSira1="Toplamda {} Adet üründen {} Adeti Yüksek Fiyatlı".format(hucreurunsay,altSira)

        yazOku.append({'A': altSira1})
        break

    else :

        hucreIci=oku['A{}'.format(hucreNo)].value

        urunAdresi=hucreIci

        baslik = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/50.0.2661.102 Safari/537.36'}

        toplama = requests.get(urunAdresi, headers=baslik)

        if toplama.status_code == 200:

                    secici = Selector(toplama.text)

                    duzenle = re.compile('<.*?>|&([a-z0-9]+|#[0-9]{1,6}|#x[0-9a-f]{1,6});')

                    urun = secici.css('#product-detail-app > div > div.pr-cn > div.pr-cn-in > div.pr-in-w > div:nth-child(1) > div.pr-in-cn > h1').get()

                    satici = secici.css('#product-detail-app > div > div.pr-cn > div.pr-cn-in > div.pr-in-at > div.pr-in-sl-cnt > div > div.sl-nm > a').get()

                    pfiyat = secici.xpath('/html/body/div[3]/div/div/div[2]/div[2]/div[1]/div[1]/div[1]/div[2]/div/div/span[1]').get()

                    sfiyat = secici.xpath('/html/body/div[3]/div/div/div[2]/div[2]/div[1]/div[1]/div[1]/div[2]/div/div/span[2]').get()

                    #print(satici) #Alttaki if için satıcı dönüş verisi 1 kez çalıştırılıp Alınması yeterli


                    if satici== "":

                        hucreNo += 1
                        continue



                    else:
                        print("Kontrol Edilen Ürün  : ", hucreNo)

                        print("Ürün Adı             : ", duzenle.sub('', urun).strip())

                        urunEkle=duzenle.sub('', urun).strip()

                        yazOku.append({'A':urunEkle})

                        try:
                            print("Ürün Satıcısı        : ",duzenle.sub('',satici).strip())

                            saticiEkle=duzenle.sub('', satici).strip()

                            yazOku.append({'A': saticiEkle})
                        except:

                            print("------------Satışta Değil------------")

                            yazOku.append({'A': "------------Satışta Değil------------"})

                        try:

                            print("Ürün Piyasa Fiyatı   : ", duzenle.sub('', pfiyat).strip())

                            pfiyatEkle=duzenle.sub('', pfiyat).strip()

                            yazOku.append({'A':pfiyatEkle })

                        except:
                            print("Ürün Piyasa Fiyatı   : Kampanya Uygullanmış")

                            yazOku.append({'A': "Kampanya Uygullanmış"})

                        try :

                            print("Ürün Satış  Fiyatı   : ", duzenle.sub('', sfiyat).strip())

                            sfiyatEkle=duzenle.sub('', sfiyat).strip()

                            yazOku.append({'A': sfiyatEkle})

                        except:

                            print("Ürün Satış  Fiyatı   : Ürüne İndirim Uygulanmamış")

                            yazOku.append({'A': "Ürüne İndirim Uygulanmamış"})

                            pass

                        print("Ürün Adresi          :", urunAdresi)

                        yazOku.append({'A': urunAdresi})

                        yazOku.append({'a':"*"*15})

                        print("*" * 25)
                        altSira+=1


        else:

            print('Bağlantı kurulamadı! HTTP Kodu: ', toplama.status_code)

        hucreNo+=1


yazAc.save('trendyolKontrol.xlsx')
ac.save('trendurunler.xlsx')

yazAc.close()
ac.close()
